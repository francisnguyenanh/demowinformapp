import re
import pandas as pd
import json
import datetime
from openpyxl import load_workbook

# Global variables for system id, date, and SEQ per sheet
# Lấy systemid_value từ input, nếu không nhập thì lấy mặc định
now = datetime.datetime.now()
default_systemid_value = f"{now.hour:02d}{now.minute:02d}{now.second:02d}"
try:
    user_input = input(f"Nhập systemid_value (Enter để dùng mặc định {default_systemid_value}): ")
except Exception:
    user_input = ''
systemid_value = user_input.strip() if user_input.strip() else default_systemid_value
system_date_value = now.strftime('%Y-%m-%d')
# seq_per_sheet_dict: {sheet_index: SEQ}
seq_per_sheet_dict = {}

# Mapping for MAPPING value (Excel cell value -> mapped number)
MAPPING_VALUE_DICT = {
    '項目定義書_帳票': '2',
    '項目定義書_画面': '1',
    '項目定義書_CSV': '5',
    '項目定義書_IPO図': '4',
    '項目定義書_ﾒﾆｭｰ': '3',
}

# Mapping for KOUMOKU types
KOUMOKU_TYPE_MAPPING = {
    'ラベル': '101',
    'タイトルラベル': '102',
    'テキストボックス': '103',
    'コンボボックス': '104',
    'ラジオボタン': '105',
    'チェックボックス': '106',
    'チェックリスト': '107',
    'ボタン': '108',
    'ボタングループ': '109',
    'メニュートゥール': '110',
    'メニューツリー': '111',
    '画像': '112'
}

KOUMOKU_TYPE_MAPPING_RE = {
    'ラベル': '101',
    'タイトルラベル': '102',
    'テキストボックス': '103',
    'コンボボックス': '106',
    'ラジオボタン': '107',
    'チェックボックス': '108',
    'チェックリスト': '114',
    'ボタン': '115',
    '画像': '116'
}

# Global stop values for process_koumoku_data (excluding '【項目定義】')
STOP_VALUES = {
    '【帳票データ】',
    '【ファンクション定義】',
    '【メッセージ定義】',
    '【タブインデックス定義】',
    '【CSVデータ】',
    '【備考】',
    '【運用上の注意点】',
    '【項目定義】', 
    '【一覧定義】',
    '【表示位置定義】'
}

def read_table_info(filename):
    """
    Reads the JSON content from the given filename and returns it as a dictionary.
    """
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def get_cell_value_with_merged(ws, cell_ref):
    """Helper function to get cell value considering merged cells"""
    cell = ws[cell_ref]
    if cell.value is not None:
        return cell.value
    # If cell is empty, check merged cells
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws[merged_range.start_cell.coordinate].value
    return None

def is_merged_from_to(ws, row, col_start, col_end):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_col == col_start
            and merged_range.max_col == col_end
            and merged_range.min_row <= row <= merged_range.max_row
        ):
            return True
    return False


def should_stop_logic_row(ws, check_row, stop_values, cell_b_value=None):
    if check_row > ws.max_row:
        return 'stop'
    cell_b_check = ws[f"B{check_row}"].value
    if cell_b_value is not None:
        if cell_b_check in stop_values and cell_b_check != cell_b_value:
            return 'stop'
    else:
        if cell_b_check in stop_values:
            return 'stop'

    merged_b_to_bn = is_merged_from_to(ws, check_row, 2, 66)
    merged_bc = is_merged_from_to(ws, check_row, 2, 3)
    
    if merged_bc:
        if cell_b_value== '画面' or cell_b_value == '番号':
            return 'skip'
        if not merged_b_to_bn:
            return 'stop'
    else:
        if not merged_b_to_bn:
            return 'skip'
        else:
            return 'continue'


def should_stop_row(ws, check_row, stop_values, cell_b_value=None):
    """
    Returns True if the row should stop processing for T_KIHON_PJ_KOUMOKU_RE (and similar tables):
    1. If cell B is in stop_values (excluding cell_b_value if provided)
    2. End of sheet is handled by the caller
    """
    if check_row > ws.max_row:
        return 'stop'
    cell_b_check = ws[f"B{check_row}"].value
    if cell_b_value is not None:
        if cell_b_check in stop_values and cell_b_check != cell_b_value:
            return 'stop'
    else:
        if cell_b_check in stop_values:
            return 'stop'
    # Check merged cells
    if cell_b_value is not None:
        if cell_b_value == '【項目定義】' or cell_b_value == '【ファンクション定義】':
            merged_b_to_bn = is_merged_from_to(ws, check_row, 2, 66)
            merged_bc = is_merged_from_to(ws, check_row, 2, 3)
            if merged_bc:
                if cell_b_check== '画面' or cell_b_check == '番号':
                    return 'skip'
                elif not merged_b_to_bn:
                    return 'continue'
            else:
                if merged_b_to_bn:
                    return 'create_logic'
                else:
                    return 'skip'
        elif cell_b_value == '【メッセージ定義】':
            merged_e_to_az = is_merged_from_to(ws, check_row, 5, 52)
            merged_bd = is_merged_from_to(ws, check_row, 2, 4)

            if merged_bd:
                if cell_b_check== 'ﾒｯｾｰｼﾞ' or cell_b_check == 'ｺｰﾄﾞ':
                    return 'skip'
                if not merged_e_to_az:
                    return 'skip'
                else:
                    return 'continue'
            else:
                return 'skip'
        elif cell_b_value == '【タブインデックス定義】':
            merged_e_to_bn = is_merged_from_to(ws, check_row, 5, 66)
            merged_bd = is_merged_from_to(ws, check_row, 2, 4)

            if merged_bd:
                if cell_b_check== '定義場所':
                    return 'skip'
                if not merged_e_to_bn:
                    return 'skip'
                else:
                    return 'continue'
            else:
                return 'skip'
        elif cell_b_value == '【備考】':
            return 'stop'
        elif cell_b_value == '【帳票データ】':
            return 'stop'
        elif cell_b_value == '【CSVデータ】':
            return 'stop'
        elif cell_b_value == '【運用上の注意点】':
            return 'stop'
        elif cell_b_value == '【表示位置定義】':
            merged_e_to_bk = is_merged_from_to(ws, check_row, 5, 63)
            merged_bd = is_merged_from_to(ws, check_row, 2, 4)

            if merged_bd:
                if cell_b_check== '定義区分':
                    return 'skip'
                if not merged_e_to_bk:
                    return 'skip'
                else:
                    return 'continue'
            else:
                return 'skip'
        elif cell_b_value == '【一覧定義】':
            merged_d_to_o = is_merged_from_to(ws, check_row, 4, 15)
            merged_bc = is_merged_from_to(ws, check_row, 2, 3)

            if merged_bc:
                if cell_b_check== '画面' or cell_b_check == '番号':
                    return 'skip'
                if not merged_d_to_o:
                    return 'skip'
                else:
                    return 'continue'
            else:
                return 'skip'
        elif cell_b_value == '【メニュー定義】':
            return 'stop'
    else:
        return 'skip'

def set_value_generic(
    col_info,
    ws,
    row_num,
    sheet_seq,
    primary_seq_value,
    secondary_seq_value=None,
    seq_mappings=None,
    reference_mappings=None,
    fallback_processor=None,
    table_name=None
):
    """
    Refactored generic function to process column values for all table types.
    This version improves readability and manageability by modularizing logic.
    """
    val_rule = col_info.get('VALUE', '')
    cell_fix = col_info.get('CELL_FIX', '').strip()
    col_logic = col_info.get('CELL_LOGIC', '').strip()
    col_name = col_info.get('COLUMN_NAME', '')
    data_type = col_info.get('DATA_TYPE', '').lower()

    def get_seq_value():
        if seq_mappings and col_name in seq_mappings:
            seq_val = seq_mappings[col_name]
            return str(seq_val) if seq_val is not None else "''"
        return None

    def get_seq_reference_value():
        if reference_mappings and val_rule in reference_mappings:
            ref_val = reference_mappings[val_rule]
            return str(ref_val) if ref_val is not None else "''"
        return None

    def handle_mapping():
        mapped_val = ''
        if cell_fix:
            cell_ref = f"{cell_fix}{row_num}"
            cell_value = ws[cell_ref].value if ws[cell_ref].value else None
        else:
            cell_ref = f"{col_logic}{row_num}"
            cell_value = get_cell_value_with_merged(ws, cell_ref)
            if col_name == 'KOUMOKU_SYURUI_CD' and isinstance(cell_value, str):
                if table_name == 'T_KIHON_PJ_KOUMOKU':
                    mapped_val = KOUMOKU_TYPE_MAPPING.get(cell_value, '')
                elif table_name == 'T_KIHON_PJ_KOUMOKU_RE':
                    mapped_val = KOUMOKU_TYPE_MAPPING_RE.get(cell_value, '')

        return f"'{mapped_val}'" if mapped_val else "''"

    def handle_empty_value():
        cell_value = None
        # Try CELL_FIX first
        if cell_fix:
            try:
                cell_ref = f"{cell_fix}{row_num}"
                cell_value = get_cell_value_with_merged(ws, cell_ref)
            except Exception:
                return "''"
        # If no value from CELL_FIX, try CELL_LOGIC
        if (cell_value is None or cell_value == '') and col_logic:
            try:
                cell_ref = f"{col_logic}{row_num}"
                cell_value = get_cell_value_with_merged(ws, cell_ref)
                # Special case for YOUKEN_NO pattern extraction
                if col_name == 'YOUKEN_NO' and isinstance(cell_value, str):
                    m = re.match(r'^\(要件№([\d\-]+)\)要件ﾛｼﾞｯｸ：', cell_value)
                    if m:
                        cell_value = m.group(1)
                    else:
                        return "''"
            except Exception:
                return "''"
                
        if cell_value is None or cell_value == '':
            return "''"
        # Additional handling for specific columns in T_KIHON_PJ_KOUMOKU
        if (
            table_name == 'T_KIHON_PJ_KOUMOKU' and
            col_name in ['ZENKAKU_MOJI_SU', 'HANKAKU_MOJI_SU', 'SEISU_KETA', 'SYOUSU_KETA'] and
            cell_value == "－"
        ):
            return "NULL"
       
        # Numeric types: do not quote, return as int/float
        if data_type in ['int'] and cell_value != "NULL":
            return int(cell_value)

        # Date/time types: quote and format
        elif data_type in ['date', 'datetime', 'smalldatetime', 'datetime2', 'datetimeoffset', 'time']:
            if isinstance(cell_value, datetime.datetime):
                return f"'{cell_value.strftime('%Y-%m-%d %H:%M:%S')}'"
            elif isinstance(cell_value, str):
                return f"'{cell_value}'"
            else:
                return f"'{str(cell_value)}'"
        # NVARCHAR: N'...'
        elif data_type == 'nvarchar':
            return f"N'{cell_value}'"
        # Default: quote as string
        else:
            return f"'{cell_value}'"

    # Main logic starts here
    # Handle AUTO_ID cases with sequence mappings
    if val_rule == 'AUTO_ID':
        seq_val = get_seq_value()
        if seq_val is not None:
            return seq_val

    # Handle specific reference mappings
    ref_val = get_seq_reference_value()
    if ref_val is not None:
        return ref_val

    # Handle MAPPING case
    if val_rule == 'MAPPING':
        return handle_mapping()

    # Handle empty value rule (direct cell reading)
    if val_rule == '':
        return handle_empty_value()

    # Handle T_KIHON_PJ_GAMEN.SEQ reference
    if val_rule == 'T_KIHON_PJ_GAMEN.SEQ':
        return str(sheet_seq) if sheet_seq is not None else "''"

    # Use fallback processor if provided
    #if fallback_processor:
    #    return fallback_processor(col_info, ws, systemid_value, system_date_value)
    # Default fallback to original process_column_value
    val = "''"
    
    if val_rule == 'BLANK':
        val = "''"
    elif val_rule == 'NULL':
        val = "NULL"
    elif val_rule == 'SYSTEMID':
        val = f"'{systemid_value}'"
    elif val_rule == 'T_KIHON_PJ.SYSTEM_ID':
        val = f"'{systemid_value}'"
        
    return val if val else "''"
    #return column_value(col_info, ws, systemid_value, system_date_value)


def koumoku_set_value(col_info, ws, row_num, sheet_seq, seq_k_value, seq_k_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU table"""
    seq_mappings = {
        'SEQ_K': seq_k_value,
        'ROW_NO': seq_k_value,
        'SEQ_K_L': seq_k_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU.SEQ_K': seq_k_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_k_value,
        secondary_seq_value=seq_k_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        fallback_processor=column_value,
        table_name='T_KIHON_PJ_KOUMOKU'
    )


def func_set_value(col_info, ws, row_num, sheet_seq, seq_f_value, seq_f_l_value=None):
    """Process column value for T_KIHON_PJ_FUNC table"""
    seq_mappings = {
        'SEQ_F': seq_f_value,
        'ROW_NO': seq_f_value,
        'SEQ_F_L': seq_f_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_FUNC.SEQ_F': seq_f_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_f_value,
        secondary_seq_value=seq_f_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        fallback_processor=column_value
    )


def csv_set_value(col_info, ws, row_num, sheet_seq, seq_csv_value, seq_csv_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU_CSV table"""
    seq_mappings = {
        'SEQ_CSV': seq_csv_value,
        'ROW_NO': seq_csv_value,
        'SEQ_CSV_L': seq_csv_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU_CSV.SEQ_CSV': seq_csv_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_csv_value,
        secondary_seq_value=seq_csv_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        fallback_processor=column_value
    )


def re_set_value(col_info, ws, row_num, sheet_seq, seq_re_value, seq_re_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU_RE table"""
    seq_mappings = {
        'SEQ_RE': seq_re_value,
        'ROW_NO': seq_re_value,
        'SEQ_RE_L': seq_re_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU_RE.SEQ_RE': seq_re_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_re_value,
        secondary_seq_value=seq_re_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        fallback_processor=column_value, 
        table_name='T_KIHON_PJ_KOUMOKU_RE'
    )


def message_set_value(col_info, ws, row_num, sheet_seq, seq_ms_value):
    """Process column value for T_KIHON_PJ_MESSAGE table"""
    seq_mappings = {
        'SEQ_MS': seq_ms_value,
        'ROW_NO': seq_ms_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_ms_value,
        seq_mappings=seq_mappings,
        fallback_processor=column_value
    )


def tab_set_value(col_info, ws, row_num, sheet_seq, seq_t_value):
    """Process column value for T_KIHON_PJ_TAB table"""
    seq_mappings = {
        'SEQ_T': seq_t_value,
        'ROW_NO': seq_t_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_t_value,
        seq_mappings=seq_mappings,
        fallback_processor=column_value
    )


def ichiran_set_value(col_info, ws, row_num, sheet_seq, seq_i_value):
    """Process column value for T_KIHON_PJ_ICHIRAN table"""
    seq_mappings = {
        'SEQ_I': seq_i_value,
        'ROW_NO': seq_i_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_i_value,
        seq_mappings=seq_mappings,
        fallback_processor=column_value
    )


def menu_set_value(col_info, ws, row_num, sheet_seq, seq_m_value):
    """Process column value for T_KIHON_PJ_MENU table"""
    seq_mappings = {
        'SEQ_M': seq_m_value,
        'ROW_NO': seq_m_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_m_value,
        seq_mappings=seq_mappings,
        fallback_processor=column_value
    )


def ipo_set_value(col_info, ws, row_num, sheet_seq, seq_ipo_value):
    """Process column value for T_KIHON_PJ_IPO table"""
    seq_mappings = {
        'SEQ_IPO': seq_ipo_value,
        'ROW_NO': seq_ipo_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_ipo_value,
        seq_mappings=seq_mappings,
        fallback_processor=column_value
    )


def column_value(col_info, ws, systemid_value, system_date_value, seq_value=None, jyun_value=None):
    """Process column value based on VALUE rules"""
    val_rule = col_info.get('VALUE', '')
    cell_fix = col_info.get('CELL_FIX', '').strip()
    col_name = col_info.get('COLUMN_NAME', '')
    
    if val_rule == 'BLANK':
        val = "''"
    elif val_rule == 'NULL':
        val = "NULL"
    elif val_rule == 'SYSTEMID':
        val = f"'{systemid_value}'"
    elif val_rule == 'T_KIHON_PJ.SYSTEM_ID':
        val = f"'{systemid_value}'"
    elif val_rule == 'AUTO_ID' and col_name == 'SEQ':
        val = str(seq_value) if seq_value is not None else "''"
    elif val_rule == 'AUTO_ID' and col_name == 'JYUN':
        val = str(jyun_value) if jyun_value is not None else "''"
    elif val_rule in ('SYSTEM DATE', 'AUTO_TIME'):
        val = f"'{system_date_value}'"
    elif val_rule == 'MAPPING':
        cell_value = ws[cell_fix].value if cell_fix else None
        val = MAPPING_VALUE_DICT.get(cell_value, "''")
    elif val_rule == '':
        if cell_fix:
            try:
                cell_value = get_cell_value_with_merged(ws, cell_fix)
                if cell_value is None:
                    val = "''"
                else:
                    data_type = col_info.get('DATA_TYPE', '').lower()
                    # Numeric types: do not quote
                    if data_type in ['int', 'bigint', 'smallint', 'tinyint', 'decimal', 'numeric', 'float', 'real', 'money', 'smallmoney']:
                        try:
                            if isinstance(cell_value, (int, float)):
                                val = str(cell_value)
                            else:
                                cell_str = str(cell_value).replace(',', '').replace(' ', '')
                                if '.' in cell_str:
                                    val = str(float(cell_str))
                                else:
                                    val = str(int(cell_str))
                        except Exception:
                            val = '0'
                    # Date/time types: quote and format
                    elif data_type in ['date', 'datetime', 'smalldatetime', 'datetime2', 'datetimeoffset', 'time']:
                        if isinstance(cell_value, datetime.datetime):
                            val = f"'{cell_value.strftime('%Y-%m-%d %H:%M:%S')}'"
                        elif isinstance(cell_value, str):
                            val = f"'{cell_value}'"
                        else:
                            val = f"'{str(cell_value)}'"
                    # NVARCHAR: N'...'
                    elif data_type == 'nvarchar':
                        val = f"N'{cell_value}'"
                    # Default: quote as string
                    else:
                        val = f"'{cell_value}'"
            except Exception:
                val = "''"
        else:
            val = "''"
    else:
        # Other values, treat as string literal
        # Add N prefix for nvarchar columns
        if col_info.get('DATA_TYPE', '').lower() == 'nvarchar':
            val = f"N'{val_rule}'"
        else:
            val = f"'{val_rule}'"
    
    return val
  


def generate_insert_statements_from_excel(excel_file, sheet_index, table_info_file, table_key):
    """
    Unified function to generate INSERT statements for all table types
    """
    # Read table info JSON
    table_info = read_table_info(table_info_file)
    if table_key not in table_info:
        raise ValueError(f"Table key '{table_key}' not found in table info.")
    
    columns_info = table_info[table_key]
    insert_statements = []
    
    # Use global systemid_value and system_date_value
    global systemid_value, system_date_value
    
    if table_key == 'T_KIHON_PJ_GAMEN':
        # Special handling for T_KIHON_PJ_GAMEN: process multiple sheets
        global seq_per_sheet_dict
        wb = load_workbook(excel_file, data_only=True)
        sheetnames = wb.sheetnames
        seq_per_sheet = 1
        allowed_b2_values = set(MAPPING_VALUE_DICT.keys())
        for sheet_idx in range(2, len(sheetnames)):
            ws = wb[sheetnames[sheet_idx]]
            try:
                sheet_check_value = ws["B2"].value
            except Exception:
                sheet_check_value = None
            if sheet_check_value not in allowed_b2_values:
                continue
            row_data = {}
            seq_value = seq_per_sheet
            jyun_value = seq_value
            seq_per_sheet_dict[sheet_idx] = seq_value
            for col_info in columns_info:
                col_name = col_info.get('COLUMN_NAME', '')
                val = column_value(col_info, ws, systemid_value, system_date_value, seq_value, jyun_value)
                row_data[col_name] = val
            columns_str = ", ".join(row_data.keys())
            values_str = ", ".join(row_data.values())
            sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
            seq_per_sheet += 1
    
    elif table_key == 'T_KIHON_PJ':
        # Special handling for T_KIHON_PJ: single insert statement
        wb = load_workbook(excel_file, data_only=True)
        sheetnames = wb.sheetnames
        if sheet_index >= len(sheetnames):
            raise ValueError(f"Sheet index {sheet_index} out of range.")
        ws = wb[sheetnames[sheet_index]]
        
        cols = []
        vals = []
        for col_info in columns_info:
            col_name = col_info['COLUMN_NAME']
            cols.append(col_name)
            val = column_value(col_info, ws, systemid_value, system_date_value)
            vals.append(val)

        columns_str = ", ".join(cols)
        values_str = ", ".join(vals)
        sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
        insert_statements.append(sql)
    
    else:
        # Default handling for other tables: process each row in the sheet
        df = pd.read_excel(excel_file, sheet_name=sheet_index, engine='openpyxl')
        wb = load_workbook(excel_file, data_only=True)
        sheetnames = wb.sheetnames
        if sheet_index >= len(sheetnames):
            raise ValueError(f"Sheet index {sheet_index} out of range.")
        ws = wb[sheetnames[sheet_index]]
        
        for _, row in df.iterrows():
            cols = []
            vals = []
            for col_info in columns_info:
                col_name = col_info['COLUMN_NAME']
                cols.append(col_name)
                val = column_value(col_info, ws, systemid_value, system_date_value)
                vals.append(val)
            columns_str = ", ".join(cols)
            values_str = ", ".join(vals)
            sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
    
    return insert_statements


def all_tables_in_sequence(excel_file, table_info_file, output_file='insert_all.sql'):
    """
    Process all tables in the correct sequence:
    1. Create INSERT for T_KIHON_PJ
    2. Iterate through sheets (from sheet 3) to create INSERT for T_KIHON_PJ_GAMEN
    3. For each new SEQ, process T_KIHON_PJ_KOUMOKU
    4. For each new SEQ_K, process T_KIHON_PJ_KOUMOKU_LOGIC
    """
    global seq_per_sheet_dict
    
    all_insert_statements = []
    
    # Step 1: Create INSERT for T_KIHON_PJ (using sheet 3)
    print("Processing T_KIHON_PJ...")
    pj_inserts = generate_insert_statements_from_excel(excel_file, 2, table_info_file, 'T_KIHON_PJ')
    all_insert_statements.extend(pj_inserts)
    
    # Step 2: Process T_KIHON_PJ_GAMEN for sheets from index 2 onwards
    print("Processing T_KIHON_PJ_GAMEN...")
    table_info = read_table_info(table_info_file)
    gamen_columns_info = table_info.get('T_KIHON_PJ_GAMEN', [])
    
    wb = load_workbook(excel_file, data_only=True)
    sheetnames = wb.sheetnames
    seq_per_sheet = 1
    allowed_b2_values = set(MAPPING_VALUE_DICT.keys())
    
    for sheet_idx in range(2, len(sheetnames)):
        ws = wb[sheetnames[sheet_idx]]
        try:
            sheet_check_value = ws["B2"].value
        except Exception:
            sheet_check_value = None

        if sheet_check_value not in allowed_b2_values:
            continue

        # Always process T_KIHON_PJ_GAMEN
        row_data = {}
        seq_value = seq_per_sheet
        jyun_value = seq_value
        seq_per_sheet_dict[sheet_idx] = seq_value
        for col_info in gamen_columns_info:
            col_name = col_info.get('COLUMN_NAME', '')
            val = column_value(col_info, ws, systemid_value, system_date_value, seq_value, jyun_value)
            row_data[col_name] = val
        columns_str = ", ".join(row_data.keys())
        values_str = ", ".join(row_data.values())
        sql = f"INSERT INTO T_KIHON_PJ_GAMEN ({columns_str}) VALUES ({values_str});"
        all_insert_statements.append(sql)
        print(f"Processing sheet {sheet_idx}: {sheetnames[sheet_idx]} with SEQ {seq_value}")

        # Xử lý theo từng loại sheet_check_value
        if sheet_check_value == '項目定義書_帳票':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_KOUMOKU_RE, T_KIHON_PJ_KOUMOKU_RE_LOGIC
            re_inserts = re_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(re_inserts)
        elif sheet_check_value == '項目定義書_CSV':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_KOUMOKU_CSV, T_KIHON_PJ_KOUMOKU_CSV_LOGIC
            csv_inserts = csv_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(csv_inserts)
        elif sheet_check_value == '項目定義書_IPO図':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_IPO
            ipo_inserts = ipo_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(ipo_inserts)
        elif sheet_check_value == '項目定義書_ﾒﾆｭｰ':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_MENU
            menu_inserts = menu_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(menu_inserts)
        elif sheet_check_value == '項目定義書_画面':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_FUNC, T_KIHON_PJ_FUNC_LOGIC, T_KIHON_PJ_KOUMOKU, T_KIHON_PJ_KOUMOKU_LOGIC, T_KIHON_PJ_MESSAGE, T_KIHON_PJ_TAB, T_KIHON_PJ_ICHIRAN, T_KIHON_PJ_HYOUJI
            func_inserts = func_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(func_inserts)
            koumoku_inserts = koumoku_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(koumoku_inserts)
            message_inserts = message_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(message_inserts)
            tab_inserts = tab_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(tab_inserts)
            ichiran_inserts = ichiran_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(ichiran_inserts)
            hyouji_inserts = hyouji_row(
                excel_file, sheet_idx, seq_value, table_info_file
            )
            all_insert_statements.extend(hyouji_inserts)
        seq_per_sheet += 1
    
    # Write all statements to file
    with open(output_file, 'w', encoding='utf-8') as f:
        for stmt in all_insert_statements:
            f.write(stmt + '\n')
    
    print(f"All INSERT statements written to {output_file}")
    return all_insert_statements


def gen_row_single_sheet(
    excel_file,
    sheet_idx,
    sheet_seq,
    table_info_file,
    table_name,
    logic_table_name=None,
    cell_b_value='【項目定義】',
    column_value_processor=None,
    logic_processor=None,
    seq_prefix='SEQ',
    stop_values=None
):
    """
    Generic function to process table data for a single sheet
    Returns list of INSERT statements for main table and optional logic table
    """
    if stop_values is None:
        stop_values = STOP_VALUES

    table_info = read_table_info(table_info_file)
    columns_info = table_info.get(table_name, [])
    logic_columns_info = table_info.get(logic_table_name, []) if logic_table_name else []

    wb = load_workbook(excel_file, data_only=True)
    sheetnames = wb.sheetnames

    if sheet_idx >= len(sheetnames):
        return []

    ws = wb[sheetnames[sheet_idx]]
    insert_statements = []
    seq_counter = 1

    print(f"  Processing {table_name} data for sheet {sheet_idx}: {sheetnames[sheet_idx]}")
    
    # Scan from top to bottom for cell_b_value
    for row_num in range(1, ws.max_row + 1):
        cell_b = ws[f"B{row_num}"]
        if cell_b.value == cell_b_value:
            check_row = row_num + 1
            while check_row <= ws.max_row:
                should_stop = should_stop_row(ws, check_row, stop_values, cell_b_value)
                if should_stop == 'stop':
                    return insert_statements
                elif should_stop == 'skip':
                    check_row += 1
                    continue
                elif should_stop == 'continue':
                    current_seq = seq_counter
                    row_data = {}
                    for col_info in columns_info:
                        col_name = col_info.get('COLUMN_NAME', '')
                        if column_value_processor:
                            val = column_value_processor(col_info, ws, check_row, sheet_seq, current_seq)
                        else:
                            val = column_value(col_info, ws, systemid_value, system_date_value)
                        row_data[col_name] = val
                    columns_str = ", ".join(row_data.keys())
                    def sql_value(v):
                        if isinstance(v, bool):
                            return 'True' if v else 'False'
                        return str(v)
                    values_str = ", ".join(sql_value(v) for v in row_data.values())
                    sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
                    insert_statements.append(sql)
                    seq_counter += 1
                    print(f"    Created {table_name.split('_')[-1]} with {seq_prefix} {current_seq} at row {check_row}")
                    check_row += 1
                    continue
                elif should_stop == 'create_logic' and logic_table_name and logic_processor:
                    # Process logic table if provided and logic processor available
                    logic_inserts, check_row_update = logic_processor(
                        ws, check_row, sheet_seq, current_seq, logic_columns_info
                    )
                    check_row = check_row_update if check_row_update > check_row else check_row + 1
                    insert_statements.extend(logic_inserts)
                    continue
                else:
                    check_row += 1
    return insert_statements


def koumoku_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【項目定義】'
):
    """
    Process KOUMOKU data for a single sheet
    Returns list of INSERT statements for both T_KIHON_PJ_KOUMOKU and T_KIHON_PJ_KOUMOKU_LOGIC
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_KOUMOKU',
        logic_table_name='T_KIHON_PJ_KOUMOKU_LOGIC',
        cell_b_value=cell_b_value,
        column_value_processor=koumoku_set_value,
        logic_processor=koumoku_logic,
        seq_prefix='SEQ_K',
        stop_values=stop_values
    )


def logic_data_generic(
    ws, 
    start_row, 
    sheet_seq, 
    parent_seq_value, 
    logic_columns_info,
    table_name,
    column_value_processor,
    seq_counter_name,
    stop_values=None,
    cell_b_value=None
):
    """
    Generic function to process logic table data
    """
    if stop_values is None:
        stop_values = STOP_VALUES
        
    insert_statements = []
    seq_counter = 1
    
    for check_row in range(start_row, ws.max_row + 1):
        # Use appropriate stopping condition
        should_stop = should_stop_logic_row(ws, check_row, stop_values, cell_b_value)
        if should_stop == 'stop':
            return insert_statements, check_row
        elif should_stop == 'skip':
            continue
        elif should_stop == 'continue':
            # Create LOGIC insert
            row_data = {}
            for col_info in logic_columns_info:
                col_name = col_info.get('COLUMN_NAME', '')
                val = column_value_processor(col_info, ws, check_row, sheet_seq, parent_seq_value, seq_counter)
                row_data[col_name] = val
            # Nếu row_data[YOUKEN_NO] khác rỗng và có key YOUKEN_LOGIC thì set YOUKEN_LOGIC = rỗng
            if 'YOUKEN_NO' in row_data and row_data['YOUKEN_NO'] not in [None, '', "''"] and 'YOUKEN_LOGIC' in row_data:
                row_data['YOUKEN_LOGIC'] = "''"
            columns_str = ", ".join(row_data.keys())
            values_str = ", ".join(row_data.values())
            sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
            logic_type = table_name.split('_')[-1]  # Extract LOGIC type name
            print(f"      Created {logic_type} with {seq_counter_name} {seq_counter} at row {check_row}")
            seq_counter += 1
    

def koumoku_logic(ws, start_row, sheet_seq, seq_k_value, koumoku_logic_columns_info):
    """
    Process T_KIHON_PJ_KOUMOKU_LOGIC for a specific SEQ_K
    """
    return logic_data_generic(
        ws=ws,
        start_row=start_row,
        sheet_seq=sheet_seq,
        parent_seq_value=seq_k_value,
        logic_columns_info=koumoku_logic_columns_info,
        table_name='T_KIHON_PJ_KOUMOKU_LOGIC',
        column_value_processor=koumoku_set_value,
        seq_counter_name='SEQ_K_L',
        cell_b_value='【項目定義】'
    )


def func_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【ファンクション定義】'
):
    """
    Process FUNC data for a single sheet
    Returns list of INSERT statements for both T_KIHON_PJ_FUNC and T_KIHON_PJ_FUNC_LOGIC
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_FUNC',
        logic_table_name='T_KIHON_PJ_FUNC_LOGIC',
        cell_b_value=cell_b_value,
        column_value_processor=func_set_value,
        logic_processor=func_logic,
        seq_prefix='SEQ_F',
        stop_values=stop_values
    )


def func_logic(ws, start_row, sheet_seq, seq_f_value, func_logic_columns_info):
    """
    Process T_KIHON_PJ_FUNC_LOGIC for a specific SEQ_F
    """
    return logic_data_generic(
        ws=ws,
        start_row=start_row,
        sheet_seq=sheet_seq,
        parent_seq_value=seq_f_value,
        logic_columns_info=func_logic_columns_info,
        table_name='T_KIHON_PJ_FUNC_LOGIC',
        column_value_processor=func_set_value,
        seq_counter_name='SEQ_F_L',
        cell_b_value='【ファンクション定義】'
    )


def csv_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【CSVデータ】'
):
    """
    Process CSV data for a single sheet
    Returns list of INSERT statements for both T_KIHON_PJ_KOUMOKU_CSV and T_KIHON_PJ_KOUMOKU_CSV_LOGIC
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_KOUMOKU_CSV',
        logic_table_name='T_KIHON_PJ_KOUMOKU_CSV_LOGIC',
        cell_b_value=cell_b_value,
        column_value_processor=csv_set_value,
        logic_processor=csv_logic,
        seq_prefix='SEQ_CSV',
        stop_values=stop_values
    )


def csv_logic(ws, start_row, sheet_seq, seq_csv_value, csv_logic_columns_info):
    """
    Process T_KIHON_PJ_KOUMOKU_CSV_LOGIC for a specific SEQ_CSV
    """
    return logic_data_generic(
        ws=ws,
        start_row=start_row,
        sheet_seq=sheet_seq,
        parent_seq_value=seq_csv_value,
        logic_columns_info=csv_logic_columns_info,
        table_name='T_KIHON_PJ_KOUMOKU_CSV_LOGIC',
        column_value_processor=csv_set_value,
        seq_counter_name='SEQ_CSV_L',
        cell_b_value='【CSVデータ】'
    )


def re_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【項目定義】'
):
    """
    Process RE data for a single sheet
    Returns list of INSERT statements for both T_KIHON_PJ_KOUMOKU_RE and T_KIHON_PJ_KOUMOKU_RE_LOGIC
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_KOUMOKU_RE',
        logic_table_name='T_KIHON_PJ_KOUMOKU_RE_LOGIC',
        cell_b_value=cell_b_value,
        column_value_processor=lambda col_info, ws, check_row, sheet_seq, current_seq: re_set_value(col_info, ws, check_row, sheet_seq, current_seq, cell_b_value),
        seq_prefix='SEQ_RE',
        stop_values=stop_values
    )


def re_logic(ws, start_row, sheet_seq, seq_re_value, re_logic_columns_info, cell_b_value='【項目定義】'):
    """
    Process T_KIHON_PJ_KOUMOKU_RE_LOGIC for a specific SEQ_RE
    """
    return logic_data_generic(
        ws=ws,
        start_row=start_row,
        sheet_seq=sheet_seq,
        parent_seq_value=seq_re_value,
        logic_columns_info=re_logic_columns_info,
        table_name='T_KIHON_PJ_KOUMOKU_RE_LOGIC',
        column_value_processor=re_set_value,
        seq_counter_name='SEQ_RE_L',
        cell_b_value=cell_b_value
    )


def message_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【メッセージ定義】'
):
    """
    Process MESSAGE data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_MESSAGE
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_MESSAGE',
        cell_b_value=cell_b_value,
        column_value_processor=message_set_value,
        seq_prefix='SEQ_MS',
        stop_values=stop_values
    )


def tab_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【タブインデックス定義】'
):
    """
    Process TAB data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_TAB
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_TAB',
        cell_b_value=cell_b_value,
        column_value_processor=tab_set_value,
        seq_prefix='SEQ_T',
        stop_values=stop_values
    )


def hyouji_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【表示位置定義】'
):
    """
    Process HYOUJI data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_HYOUJI
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_HYOUJI',
        cell_b_value=cell_b_value,
        column_value_processor=message_set_value,  # Sử dụng processor message
        seq_prefix='SEQ_HYOUJI',
        stop_values=stop_values
    )


def ichiran_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【一覧定義】'
):
    """
    Process ICHIRAN data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_ICHIRAN
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_ICHIRAN',
        cell_b_value=cell_b_value,
        column_value_processor=ichiran_set_value,
        seq_prefix='SEQ_I',
        stop_values=stop_values
    )


def menu_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【メニュー定義】'
):
    """
    Process MENU data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_MENU
    """
    return gen_row_single_sheet(
        excel_file=excel_file,
        sheet_idx=sheet_idx,
        sheet_seq=sheet_seq,
        table_info_file=table_info_file,
        table_name='T_KIHON_PJ_MENU',
        cell_b_value=cell_b_value,
        column_value_processor=menu_set_value,
        seq_prefix='SEQ_M',
        stop_values=stop_values
    )


def ipo_row(
    excel_file, 
    sheet_idx, 
    sheet_seq, 
    table_info_file,
    stop_values=None,
    cell_b_value='【IPO定義】'
):
    """
    Process IPO data for a single sheet
    Returns list of INSERT statements for T_KIHON_PJ_IPO
    """
    if stop_values is None:
        stop_values = STOP_VALUES

    table_info = read_table_info(table_info_file)
    ipo_columns_info = table_info.get('T_KIHON_PJ_IPO', [])

    wb = load_workbook(excel_file, data_only=True)
    sheetnames = wb.sheetnames

    if sheet_idx >= len(sheetnames):
        return []

    ws = wb[sheetnames[sheet_idx]]
    insert_statements = []
    seq_ipo_counter = 1

    print(f"  Processing IPO data for sheet {sheet_idx}: {sheetnames[sheet_idx]}")
    
    # Scan from top to bottom for cell_b_value
    for row_num in range(1, ws.max_row + 1):
        cell_b = ws[f"B{row_num}"]
        if cell_b.value == cell_b_value:
            # Check subsequent rows
            for check_row in range(row_num + 1, ws.max_row + 1):
                cell_b_check = ws[f"B{check_row}"].value
                # Skip if value is cell_b_value, break if in stop_values
                if cell_b_check == cell_b_value:
                    continue
                if cell_b_check in stop_values:
                    break
                
                # Check if B and C are merged and have value != '画面' and != '番号'
                merged_bc = False
                for merged_range in ws.merged_cells.ranges:
                    if f"B{check_row}" in merged_range and f"C{check_row}" in merged_range:
                        merged_bc = True
                        break
                
                if merged_bc:
                    cell_b_val = ws[f"B{check_row}"].value
                    if cell_b_val and cell_b_val != '画面' and cell_b_val != '番号':
                        # Create IPO insert
                        current_seq_ipo = seq_ipo_counter
                        
                        row_data = {}
                        for col_info in ipo_columns_info:
                            col_name = col_info.get('COLUMN_NAME', '')
                            val = ipo_set_value(col_info, ws, check_row, sheet_seq, current_seq_ipo)
                            row_data[col_name] = val
                        
                        columns_str = ", ".join(row_data.keys())
                        values_str = ", ".join(row_data.values())
                        sql = f"INSERT INTO T_KIHON_PJ_IPO ({columns_str}) VALUES ({values_str});"
                        insert_statements.append(sql)
                        
                        print(f"    Created IPO with SEQ_IPO {current_seq_ipo} at row {check_row}")
                        
                        seq_ipo_counter += 1
    
    return insert_statements


# Example usage:
if __name__ == "__main__":
    print("Starting processing all tables in sequence...")
    all_inserts = all_tables_in_sequence('doc_gamen.xlsx', 'table_info.txt', 'insert_all.sql')
    print(f"Generated {len(all_inserts)} INSERT statements in total.")


